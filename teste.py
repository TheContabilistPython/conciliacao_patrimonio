import re
import openpyxl

# Lista de palavras compostas para buscar
keywords = [
    "DEPRECIAÇÃO ACUMULADA",
    "98 -  Máquinas e Equipamentos",
    "Móveis e Utensílios",
    "Ferramentas",
    "Imóveis",
    "Veículos",
    "Computadores e periféricos",
    "Edificações e Construções",
    "Veículos Usados",
    "Aeronaves",
    "Benfeitorias em Imóveis de Terceiros",
    "Aparelhos Telefônicos",
    "Instalações Diversas",
    "Rouparias",
    "Obras de Arte",
    "Máquinas e Equipamentos Industriais",
    "Salas Comerciais",
    "Imobilizado Diversos"
]

# Dicionário ligando {palavra} a um número
palavra_para_numero = {
    "depreciacao_acumulada": 115,
    "98_-__maquinas_e_equipamentos": 116,
    "moveis_e_utensilios": 117,
    "ferramentas": 118,
    "imoveis": 119,
    "veiculos": 120,
    "computadores_e_perifericos": 121,
    "edificacoes_e_construcoes": 122,
    "veiculos_usados": 2694,
    "aeronaves": 2695,
    "benfeitorias_em_imoveis_de_terceiros": 2696,
    "aparelhos_telefonicos": 2697,
    "instalacoes_diversas": 2698,
    "rouparias": 2699,
    "obras_de_arte": 3211,
    "maquinas_e_equipamentos_industriais": 3351,
    "salas_comerciais": 3754,
    "imobilizado_diversos": 3820
}

# Função para normalizar as palavras
def normalizar_palavra(palavra):
    return palavra.lower().replace("ç", "c").replace("ã", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u").replace("ê", "e").replace("â", "a").replace("ô", "o").replace("á", "a").replace("à", "a").replace("ü", "u").replace("õ", "o").replace("ñ", "n").replace("í", "i").replace("ú", "u").replace("é", "e").replace("ó", "o").replace("ã", "a").replace("ç", "c").replace(" ", "_")

# Função para buscar palavras no arquivo HTML e capturar valores
def buscar_palavras_em_html(arquivo_html):
    with open(arquivo_html, 'r', encoding='utf-8') as f:
        linhas = f.readlines()

    # Compilar a regex para buscar palavras-chave e capturar valores
    padrao = re.compile(r'(' + '|'.join(map(re.escape, keywords)) + r').*?>([\d.,]+)</td>.*?>([\d.,]+)</td>', re.IGNORECASE)

    # Verificar cada linha
    resultados = []
    valores_dict = {}
    for i, linha in enumerate(linhas, 1):
        print(f"Verificando linha {i}: {linha.strip()}")  # Debug: print each line being checked
        correspondencia = padrao.search(linha)
        if correspondencia:
            palavra = correspondencia.group(1)
            valor1 = correspondencia.group(2)
            resultados.append(f"{palavra}: {valor1}")
            palavra_normalizada = normalizar_palavra(palavra)
            if palavra_normalizada in palavra_para_numero:
                valores_dict[palavra_para_numero[palavra_normalizada]] = valor1

    # Retornar todas as linhas encontradas
    return resultados, valores_dict

# Função para atualizar a planilha Excel
def atualizar_planilha_excel(planilha_excel, valores_dict):
    wb = openpyxl.load_workbook(planilha_excel)
    ws = wb.active
    linhas_encontradas = False

    # Calcular a soma dos valores específicos
    soma_valores = sum(float(valores_dict[numero].replace('.', '').replace(',', '.')) for numero in [116, 117, 118, 119, 120, 121, 122, 2694, 2695, 2696, 2697, 2698, 2699, 3211, 3351, 3754, 3820] if numero in valores_dict)
    valores_dict[115] = f"{soma_valores:,.2f}".replace('.', ',').replace(',', '.')

    for row in ws.iter_rows(min_row=2):
        cell_a = row[0].value  # Coluna A (índice 0)
        cell_d = row[3].value  # Coluna D (índice 3)
        cell_e = row[4].value  # Coluna E (índice 4)
        cell_g = row[6]        # Coluna G (índice 6)
        cell_h = row[7]        # Coluna H (índice 7)
        cell_i = row[8]        # Coluna I (índice 8)

        if cell_a and cell_e:
            if cell_a in valores_dict:
                print(f"Linha encontrada: {cell_a}")
                valor1 = valores_dict.get(cell_a, None)
                if valor1:
                    valor1_float = float(valor1.replace('.', '').replace(',', '.'))
                    cell_g.value = valor1_float
                    print(f"Valor encontrado em G: {cell_g.value}")
                    cell_h.value = f"=E{cell_h.row}-G{cell_h.row}"
                    linhas_encontradas = True

            # Caso específico para A = 115
            if cell_a == 115:
                cell_g.value = soma_valores
                print(f"Soma dos valores em G: {cell_g.value}")
                cell_h.value = f"=E{cell_h.row}-G{cell_h.row}"
                linhas_encontradas = True

            # Escrever "OK" na coluna I para os números especificados
            if cell_a in [3351, 120, 3350, 102]:
                cell_i.value = "OK"

    if not linhas_encontradas:
        print("Linhas não encontradas.")

    wb.save(planilha_excel)

    # Verificar se os valores foram encontrados nas linhas específicas
    for numero in [116, 115, 2696, 121]:
        for row in ws.iter_rows(min_row=2):
            cell_a = row[0].value
            if cell_a == numero:
                print(f"Valor na linha com A={numero}: E={row[4].value}, G={row[6].value}, H={row[7].value}")

# Função para procurar '(-) Computadores Periféricos' na coluna D e retornar a linha
def procurar_computadores_perifericos(planilha_excel):
    wb = openpyxl.load_workbook(planilha_excel)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        cell_d = row[3].value  # Coluna D (índice 3)
        if "(-) Computadores Periféricos" in str(cell_d):
            print(f"'(-) Computadores Periféricos' encontrado na linha: {row[0].row}")

# Exemplo de uso
arquivo_html = r"C:\Users\contabil18\Downloads\testeteste.htm"
planilha_excel = r"C:\projeto\planilhas\balancete\CONCILIACAO_13701_112024.xlsx" 

resultados, valores_dict = buscar_palavras_em_html(arquivo_html)

# Exibir resultados encontrados
if resultados:
    for resultado in resultados:
        print(resultado)
else:
    print("Nenhuma correspondência encontrada.")

# Atualizar a planilha Excel
atualizar_planilha_excel(planilha_excel, valores_dict)

# Exibir valores específicos
for codigo, valor in valores_dict.items():
    print(f"{codigo}: {valor}")

# Verificar se os valores foram encontrados nas linhas específicas
wb = openpyxl.load_workbook(planilha_excel)
ws = wb.active
for numero in [116, 115, 2696, 121]:
    for row in ws.iter_rows(min_row=2):
        cell_a = row[0].value
        if cell_a == numero:
            print(f"Valor na linha com A={numero}: E={row[4].value}, G={row[6].value}, H={row[7].value}")

# Procurar '(-) Computadores Periféricos' na coluna D
procurar_computadores_perifericos(planilha_excel)

print(valores_dict)




